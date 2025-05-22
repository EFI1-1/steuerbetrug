from fontTools.t1Lib import write
from openpyxl import Workbook, load_workbook
import os
import csv
from product import Product

class ExcelManager:
    def __init__(self, filename: str):
        self.filename = filename
        self.sheet_name = "Liste"
        self._initialize_file()

    def _initialize_file(self):
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append(["Datum", "Produktname", "Menge", "Preis pro Stk (€)", "Rabatt (%)", "Gesamt (€)", "MwSt (%)", "MwSt (€)", "Netto(€)"])
            wb.save(self.filename)

    def save_entry(self, product: Product):
        wb = load_workbook(self.filename)
        ws = wb[self.sheet_name]
        ws.append(product.file_list())
        wb.save(self.filename)
        print("In Excel gespeichert!")

    def save_csv(self, product: Product, csv_filename: str = "output.csv"):
        write_header = not os.path.exists(csv_filename)
        with open(csv_filename, mode="a", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            if write_header:
                writer.writerow(["Datum", "Produktname", "Menge", "Einzelpreis (€)", "Rabatt (%)", "Brutto (€)", "MwSt (€)", "Netto (€)"])
            writer.writerow(product.file_list())
        print("In CSV gespeichert.")